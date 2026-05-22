"""Convert questions.xlsx or questions.csv to questions.json for the PWA."""
import csv, json, sys, os

def convert_csv(csv_path, json_path):
    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        questions = []
        for row in reader:
            stem = row.get('题干', '').strip()
            if not stem:
                continue
            q = {
                'id': int(row.get('题号', 0)),
                'stem': stem,
                'A': row.get('A', '').strip(),
                'B': row.get('B', '').strip(),
                'C': row.get('C', '').strip(),
                'D': row.get('D', '').strip(),
                'E': row.get('E', '').strip(),
                'answer': row.get('答案', '').strip(),
                'difficulty': row.get('难度', '无').strip() or '无',
                'qtype': row.get('题型', '单选题').strip() or '单选题',
                'category': row.get('类别', '').strip()
            }
            questions.append(q)
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(questions, f, ensure_ascii=False)
    return len(questions)

def convert_xlsx(xlsx_path, json_path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Error: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
        print("Falling back to CSV mode. Please save as CSV or install openpyxl.", file=sys.stderr)
        return None

    wb = load_workbook(xlsx_path, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        print("Error: Excel file has no data rows", file=sys.stderr)
        sys.exit(1)

    headers = [str(h).strip() if h else '' for h in rows[0]]
    questions = []
    for row in rows[1:]:
        vals = {headers[i]: (str(row[i]).strip() if row[i] is not None else '') for i in range(len(headers))}
        stem = vals.get('题干', '')
        if not stem:
            continue
        q = {
            'id': int(vals.get('题号', 0) or 0),
            'stem': stem,
            'A': vals.get('A', ''),
            'B': vals.get('B', ''),
            'C': vals.get('C', ''),
            'D': vals.get('D', ''),
            'E': vals.get('E', ''),
            'answer': vals.get('答案', ''),
            'difficulty': vals.get('难度', '无') or '无',
            'qtype': vals.get('题型', '单选题') or '单选题',
            'category': vals.get('类别', '')
        }
        questions.append(q)

    wb.close()
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(questions, f, ensure_ascii=False)
    return len(questions)

if __name__ == '__main__':
    script_dir = os.path.dirname(os.path.abspath(__file__))

    xlsx_path = os.path.join(script_dir, 'questions.xlsx')
    csv_path = os.path.join(script_dir, 'questions.csv')
    json_path = os.path.join(script_dir, 'questions.json')

    count = None

    if os.path.exists(xlsx_path):
        count = convert_xlsx(xlsx_path, json_path)
        if count is not None:
            print(f'Converted {count} questions from questions.xlsx -> questions.json')
        else:
            # Fallback to CSV
            if os.path.exists(csv_path):
                count = convert_csv(csv_path, json_path)
                print(f'Converted {count} questions from questions.csv -> questions.json')

    elif os.path.exists(csv_path):
        count = convert_csv(csv_path, json_path)
        print(f'Converted {count} questions from questions.csv -> questions.json')
    else:
        print('Error: No questions.xlsx or questions.csv found', file=sys.stderr)
        sys.exit(1)

    if count:
        print(f'Successfully generated {count} questions')
